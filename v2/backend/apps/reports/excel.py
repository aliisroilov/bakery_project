"""openpyxl helpers — keep styling logic out of views."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(start_color="FFEAD5", end_color="FFEAD5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="7C2D12")
CENTER = Alignment(horizontal="center", vertical="center")


def make_workbook(sheet_title: str, headers: list[str], rows: list[list]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel limit

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for r, row in enumerate(rows, start=2):
        # Only write as many columns as there are headers — reports may carry
        # extra hidden trailing fields (row_type, week date ranges) that must
        # not leak into the spreadsheet.
        for c, value in enumerate(row[: len(headers)], start=1):
            ws.cell(row=r, column=c, value=value)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
